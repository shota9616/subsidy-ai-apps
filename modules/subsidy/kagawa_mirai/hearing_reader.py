"""香川県未来投資応援補助金 ヒアリングシート読み込み"""

import openpyxl

from .data_models import (
    HearingData, CompanyInfo, PriceImpact, BusinessPlan,
    EffectPlan, WagePlan, FinancialInfo, SubsidyExpenseItem,
)


def read_hearing_sheet(file_path: str) -> HearingData:
    """ヒアリングシートExcelを読み込み、HearingDataを返す。

    Args:
        file_path: ヒアリングシートのファイルパス

    Returns:
        HearingData: 読み込んだヒアリングデータ

    Raises:
        FileNotFoundError: ファイルが見つからない場合
        ValueError: シート構造が不正な場合
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    data = HearingData()

    def get_cell(sheet_name, row, col, default=""):
        try:
            ws = wb[sheet_name]
            val = ws.cell(row=row, column=col).value
            return val if val is not None else default
        except Exception:
            return default

    # Sheet 1: 企業基本情報
    data.company = CompanyInfo(
        name=str(get_cell("1_企業基本情報", 6, 2, "")),
        corporate_number=str(get_cell("1_企業基本情報", 7, 2, "")),
        representative=str(get_cell("1_企業基本情報", 8, 2, "")),
        representative_title=str(get_cell("1_企業基本情報", 9, 2, "代表取締役")),
        postal_code=str(get_cell("1_企業基本情報", 10, 2, "")),
        address=str(get_cell("1_企業基本情報", 11, 2, "")),
        phone=str(get_cell("1_企業基本情報", 12, 2, "")),
        fax=str(get_cell("1_企業基本情報", 13, 2, "")),
        email=str(get_cell("1_企業基本情報", 14, 2, "")),
        industry=str(get_cell("1_企業基本情報", 15, 2, "")),
        industry_code=str(get_cell("1_企業基本情報", 16, 2, "")),
        business_description=str(get_cell("1_企業基本情報", 17, 2, "")),
        employee_count=int(get_cell("1_企業基本情報", 18, 2, 0) or 0),
        capital=int(get_cell("1_企業基本情報", 19, 2, 0) or 0),
        established_date=str(get_cell("1_企業基本情報", 20, 2, "")),
        fiscal_month=int(get_cell("1_企業基本情報", 21, 2, 0) or 0),
        entity_type=str(get_cell("1_企業基本情報", 22, 2, "")),
        sales_category=str(get_cell("1_企業基本情報", 23, 2, "")),
    )

    # Sheet 2: 事業と物価高騰
    data.price_impact = PriceImpact(
        history=str(get_cell("2_事業と物価高騰", 4, 2, "")),
        main_business=str(get_cell("2_事業と物価高騰", 5, 2, "")),
        strengths=str(get_cell("2_事業と物価高騰", 6, 2, "")),
        customers=str(get_cell("2_事業と物価高騰", 7, 2, "")),
        achievements=str(get_cell("2_事業と物価高騰", 8, 2, "")),
        material_name=str(get_cell("2_事業と物価高騰", 10, 2, "")),
        price_increase_rate=str(get_cell("2_事業と物価高騰", 11, 2, "")),
        monthly_cost_increase=str(get_cell("2_事業と物価高騰", 12, 2, "")),
        energy_impact=str(get_cell("2_事業と物価高騰", 13, 2, "")),
        labor_cost_impact=str(get_cell("2_事業と物価高騰", 14, 2, "")),
        annual_total_increase=str(get_cell("2_事業と物価高騰", 15, 2, "")),
        cost_to_sales_ratio=str(get_cell("2_事業と物価高騰", 16, 2, "")),
        countermeasures=str(get_cell("2_事業と物価高騰", 17, 2, "")),
        limitations=str(get_cell("2_事業と物価高騰", 18, 2, "")),
    )

    # Sheet 3: 補助事業の内容
    data.business = BusinessPlan(
        project_name=str(get_cell("3_補助事業の内容", 4, 2, "")),
        purpose=str(get_cell("3_補助事業の内容", 5, 2, "")),
        method=str(get_cell("3_補助事業の内容", 6, 2, "")),
        current_field=str(get_cell("3_補助事業の内容", 7, 2, "")),
        plan_field=str(get_cell("3_補助事業の内容", 8, 2, "")),
        equipment_name=str(get_cell("3_補助事業の内容", 10, 2, "")),
        equipment_description=str(get_cell("3_補助事業の内容", 11, 2, "")),
        equipment_maker=str(get_cell("3_補助事業の内容", 12, 2, "")),
        selection_reason=str(get_cell("3_補助事業の内容", 13, 2, "")),
        before_process=str(get_cell("3_補助事業の内容", 14, 2, "")),
        after_process=str(get_cell("3_補助事業の内容", 15, 2, "")),
        comparison=str(get_cell("3_補助事業の内容", 16, 2, "")),
        schedule_order=str(get_cell("3_補助事業の内容", 18, 2, "")),
        schedule_delivery=str(get_cell("3_補助事業の内容", 19, 2, "")),
        schedule_start=str(get_cell("3_補助事業の内容", 20, 2, "")),
        schedule_complete=str(get_cell("3_補助事業の内容", 21, 2, "")),
    )

    # Sheet 4: 補助事業の効果
    data.effect = EffectPlan(
        sales_increase_annual=int(get_cell("4_補助事業の効果", 4, 2, 0) or 0),
        sales_increase_reason=str(get_cell("4_補助事業の効果", 5, 2, "")),
        cost_reduction_annual=int(get_cell("4_補助事業の効果", 6, 2, 0) or 0),
        cost_reduction_reason=str(get_cell("4_補助事業の効果", 7, 2, "")),
        useful_life=str(get_cell("4_補助事業の効果", 9, 2, "")),
        maintenance=str(get_cell("4_補助事業の効果", 10, 2, "")),
        ease_of_operation=str(get_cell("4_補助事業の効果", 11, 2, "")),
        payback_estimate=str(get_cell("4_補助事業の効果", 12, 2, "")),
        regional_contribution=str(get_cell("4_補助事業の効果", 14, 2, "")),
        reference_for_others=str(get_cell("4_補助事業の効果", 15, 2, "")),
        other_notes=str(get_cell("4_補助事業の効果", 16, 2, "")),
    )

    # Sheet 5: 賃上げ計画
    data.wage = WagePlan(
        start_date=str(get_cell("5_賃上げ計画", 4, 2, "")),
        target_employees=str(get_cell("5_賃上げ計画", 5, 2, "")),
        method=str(get_cell("5_賃上げ計画", 6, 2, "")),
        amount=str(get_cell("5_賃上げ計画", 7, 2, "")),
        annual_increase=int(get_cell("5_賃上げ計画", 8, 2, 0) or 0),
        funding_source=str(get_cell("5_賃上げ計画", 9, 2, "")),
    )

    # Sheet 6: 財務情報
    data.financial = FinancialInfo(
        sales=int(get_cell("6_財務情報", 5, 2, 0) or 0),
        operating_profit=int(get_cell("6_財務情報", 6, 2, 0) or 0),
        personnel_cost=int(get_cell("6_財務情報", 7, 2, 0) or 0),
        depreciation=int(get_cell("6_財務情報", 8, 2, 0) or 0),
        salary_total=int(get_cell("6_財務情報", 9, 2, 0) or 0),
        employee_count=int(get_cell("6_財務情報", 10, 2, 1) or 1),
    )

    # Sheet 7: 補助対象経費
    try:
        ws_exp = wb["7_補助対象経費"]
        for row in range(5, 20):
            item_name = ws_exp.cell(row=row, column=3).value
            if not item_name:
                continue
            data.expenses.append(SubsidyExpenseItem(
                category=str(ws_exp.cell(row=row, column=2).value or ""),
                item_name=str(item_name),
                amount=int(ws_exp.cell(row=row, column=4).value or 0),
                has_quote=bool(ws_exp.cell(row=row, column=5).value),
                note=str(ws_exp.cell(row=row, column=6).value or ""),
            ))
    except KeyError:
        pass

    return data


def hearing_to_prompt_data(data: HearingData) -> dict:
    """HearingDataをgenerate_plan.pyのプロンプト用dictに変換"""
    return {
        "company": {
            "company_name": data.company.name,
            "industry": data.company.industry,
            "business_description": data.company.business_description,
            "employee_count": f"{data.company.employee_count}名",
            "established_date": data.company.established_date,
            "capital": f"{data.company.capital:,}千円" if data.company.capital else "",
            "history": data.price_impact.history,
            "strengths": data.price_impact.strengths,
            "customers": data.price_impact.customers,
            "achievements": data.price_impact.achievements,
        },
        "price_impact": {
            "material_name": data.price_impact.material_name,
            "price_increase_rate": data.price_impact.price_increase_rate,
            "monthly_cost_increase": data.price_impact.monthly_cost_increase,
            "energy_impact": data.price_impact.energy_impact,
            "labor_cost_impact": data.price_impact.labor_cost_impact,
            "annual_total_increase": data.price_impact.annual_total_increase,
            "cost_to_sales_ratio": data.price_impact.cost_to_sales_ratio,
            "countermeasures": data.price_impact.countermeasures,
            "limitations": data.price_impact.limitations,
        },
        "business": {
            "project_name": data.business.project_name,
            "purpose": data.business.purpose,
            "method": data.business.method,
            "equipment_name": data.business.equipment_name,
            "equipment_description": data.business.equipment_description,
            "selection_reason": data.business.selection_reason,
            "before_process": data.business.before_process,
            "after_process": data.business.after_process,
            "schedule_order": data.business.schedule_order,
            "schedule_delivery": data.business.schedule_delivery,
            "schedule_start": data.business.schedule_start,
            "schedule_complete": data.business.schedule_complete,
        },
        "effect": {
            "sales_increase_annual": f"{data.effect.sales_increase_annual:,}円",
            "sales_increase_reason": data.effect.sales_increase_reason,
            "cost_reduction_annual": f"{data.effect.cost_reduction_annual:,}円",
            "cost_reduction_reason": data.effect.cost_reduction_reason,
            "useful_life": data.effect.useful_life,
            "maintenance": data.effect.maintenance,
            "ease_of_operation": data.effect.ease_of_operation,
            "payback_period": data.effect.payback_estimate,
            "regional_contribution": data.effect.regional_contribution,
            "reference_for_others": data.effect.reference_for_others,
            "other_notes": data.effect.other_notes,
        },
        "wage": {
            "start_date": data.wage.start_date,
            "target_employees": data.wage.target_employees,
            "method": data.wage.method,
            "amount": data.wage.amount,
            "annual_increase": f"{data.wage.annual_increase:,}円",
            "funding_source": data.wage.funding_source,
        },
    }
